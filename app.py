from flask import Flask, render_template, request, redirect, url_for
import openpyxl

app = Flask(__name__)
from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import load_workbook
import smtplib

# Email credentials
SMTP_SERVER = 'smtp.example.com'
SMTP_PORT = 587
EMAIL_USER = 'your_email@example.com'
EMAIL_PASS = 'your_email_password'

def read_excel():
    wb = load_workbook("Notification.xlsx")
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "name": row[0],
            "email": row[1],
            "attendance": row[2],
            "status": row[3]
        })
    return data


def send_email(recipient_email, subject, message):
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASS)
            email_message = f"Subject: {subject}\n\n{message}"
            server.sendmail(EMAIL_USER, recipient_email, email_message)
            print(f"Notification sent to {recipient_email}")
    except Exception as e:
        print(f"Error sending email to {recipient_email}: {e}")
